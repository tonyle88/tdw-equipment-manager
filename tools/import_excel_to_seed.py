from __future__ import annotations

import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


GROUP_MAP = {
    "MÁY TÍNH - LAPTOP": ("MAY_TINH_LAPTOP", "Máy tính - Laptop"),
    "MÁY TÍNH VẬN HÀNH HỆ THỐNG SCADA - LOGGER- DATA TDW": (
        "SCADA_LOGGER_DATA",
        "SCADA - Logger - Data TDW",
    ),
    "Ổ CỨNG - THIẾT BỊ ĐIỆN - ĐIỆN TỬ": (
        "O_CUNG_THIET_BI_DIEN_TU",
        "Ổ cứng - Thiết bị điện tử",
    ),
    "MÁY IN - PHOTOCOPY - MÁY CHIẾU - TV - ĐIỆN THOẠI": (
        "MAY_IN_PHOTOCOPY_MAY_CHIEU_TV_DIEN_THOAI",
        "Máy in - Photocopy - Máy chiếu - TV - Điện thoại",
    ),
    "THIẾT BỊ LƯU KHO - KÉM PHẨM CHẤT": (
        "LUU_KHO_KEM_PHAM_CHAT",
        "Thiết bị lưu kho - Kém phẩm chất",
    ),
}

HEADERS = [
    "asset_id",
    "asset_code",
    "asset_name",
    "asset_group",
    "asset_group_label",
    "asset_type",
    "brand",
    "model",
    "serial_number",
    "purchase_year",
    "purchase_date",
    "quantity",
    "unit_price",
    "total_price",
    "assigned_to",
    "department",
    "location",
    "software_license",
    "status",
    "quality_level",
    "warranty_until",
    "last_maintenance_date",
    "next_check_date",
    "note",
    "source_row",
    "created_at",
    "updated_at",
]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.replace("\n", " ")).strip()
    return value


def normalize_status(text):
    raw = str(text or "").upper()
    if "MẤT PHẨM" in raw or "KHÔNG SỬ DỤNG" in raw:
        return "KHONG_SU_DUNG"
    if "LƯU KHO" in raw or "THANH LÝ" in raw:
        return "LUU_KHO_THANH_LY"
    if "KÉM PHẨM" in raw:
        return "KEM_PHAM_CHAT"
    if "MỚI" in raw or "100%" in raw:
        return "MOI_100"
    if "CÒN SỬ DỤNG" in raw:
        return "CON_SU_DUNG"
    return "CAN_KIEM_TRA" if raw else ""


def normalize_quality(text):
    raw = str(text or "").upper()
    values = []
    if "MẤT PHẨM" in raw:
        values.append("Mất phẩm chất")
    if "KÉM PHẨM" in raw:
        values.append("Kém phẩm chất")
    if "MỚI" in raw or "100%" in raw:
        values.append("Mới 100%")
    if "CÒN SỬ DỤNG" in raw:
        values.append("Còn sử dụng")
    return ", ".join(dict.fromkeys(values))


def guess_asset_type(name, group_code):
    text = str(name or "").upper()
    if group_code == "MAY_TINH_LAPTOP":
        if "MÁY TÍNH ĐỂ BÀN" in text or "MÁY ĐỂ BÀN" in text or "PC " in text or text.startswith("PC"):
            return "Desktop PC"
        return "Laptop"
    if group_code == "SCADA_LOGGER_DATA" or "SERVER" in text:
        return "Server/SCADA"
    if "LAPTOP" in text or "NOTEBOOK" in text or "VIVOBOOK" in text or "THINKBOOK" in text:
        return "Laptop"
    if "MÁY TÍNH ĐỂ BÀN" in text or "PC " in text or text.startswith("PC"):
        return "Desktop PC"
    if "MÁY IN" in text or "PRINTER" in text:
        return "Máy in"
    if "PHOTOCOPY" in text:
        return "Photocopy"
    if "MÁY CHIẾU" in text:
        return "Máy chiếu"
    if "TV" in text:
        return "TV"
    if "Ổ CỨNG" in text or "SSD" in text or "HDD" in text:
        return "Ổ cứng"
    if "ĐIỆN THOẠI" in text:
        return "Điện thoại"
    if "MÀN HÌNH" in text:
        return "Màn hình"
    return "Thiết bị"


def guess_brand(name):
    brands = [
        "Dell",
        "Asus",
        "Lenovo",
        "HP",
        "Canon",
        "Epson",
        "Sony",
        "Samsung",
        "Nokia",
        "Brother",
        "WD",
        "Godex",
        "Poly",
    ]
    text = str(name or "").upper()
    for brand in brands:
        if brand.upper() in text:
            return brand
    return ""


def group_prefix(group_code):
    return {
        "MAY_TINH_LAPTOP": "LAP",
        "SCADA_LOGGER_DATA": "SCA",
        "O_CUNG_THIET_BI_DIEN_TU": "DEV",
        "MAY_IN_PHOTOCOPY_MAY_CHIEU_TV_DIEN_THOAI": "PRN",
        "LUU_KHO_KEM_PHAM_CHAT": "STO",
    }.get(group_code, "AST")


def read_assets(xlsx_path):
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    current_group = None
    counters = {}
    assets = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [clean(value) for value in row]
        non_empty = [value for value in cells if value != ""]
        if not non_empty:
            continue

        first = str(non_empty[0]).strip()
        if first in GROUP_MAP:
            current_group = GROUP_MAP[first]
            counters.setdefault(current_group[0], 0)
            continue

        if not current_group:
            continue

        if not isinstance(cells[1] if len(cells) > 1 else "", str) and not cells[1]:
            continue

        stt = cells[1] if len(cells) > 1 else ""
        if not str(stt).isdigit():
            continue

        group_code, group_label = current_group
        counters[group_code] += 1
        sequence = counters[group_code]

        name = cells[2] if len(cells) > 2 else ""
        purchase_year = cells[3] if len(cells) > 3 else ""
        quantity = cells[4] if len(cells) > 4 else ""
        assigned_to = cells[5] if len(cells) > 5 else ""
        total_price = cells[6] if len(cells) > 6 and isinstance(cells[6], (int, float)) else ""
        software = cells[7] if len(cells) > 7 else ""
        status_text = cells[8] if len(cells) > 8 else ""
        note = cells[9] if len(cells) > 9 else ""

        if group_code != "MAY_TINH_LAPTOP" and software and not any(
            key in str(software).upper() for key in ["WINDOW", "OFFICE", "LEMON", "LINUX", "SERVER"]
        ):
            note = " | ".join([str(value) for value in [software, status_text, note] if value])
            software = ""
            status_text = cells[7] if len(cells) > 7 else status_text

        asset_id = f"asset_{group_prefix(group_code).lower()}_{sequence:03d}"
        code_year = purchase_year if str(purchase_year).isdigit() else "0000"

        assets.append(
            {
                "asset_id": asset_id,
                "asset_code": f"TDW-{group_prefix(group_code)}-{code_year}-{sequence:03d}",
                "asset_name": name,
                "asset_group": group_code,
                "asset_group_label": group_label,
                "asset_type": guess_asset_type(name, group_code),
                "brand": guess_brand(name),
                "model": "",
                "serial_number": "",
                "purchase_year": purchase_year,
                "purchase_date": "",
                "quantity": quantity,
                "unit_price": "",
                "total_price": total_price,
                "assigned_to": assigned_to,
                "department": assigned_to if str(assigned_to).lower().startswith(("phòng", "p.", "p ")) else "",
                "location": "",
                "software_license": software,
                "status": normalize_status(" ".join([str(status_text), str(note)])),
                "quality_level": normalize_quality(" ".join([str(status_text), str(note)])),
                "warranty_until": "",
                "last_maintenance_date": "",
                "next_check_date": "",
                "note": " | ".join([str(value) for value in [status_text, note] if value]),
                "source_row": row_index,
                "created_at": now,
                "updated_at": now,
            }
        )

    return assets


def write_outputs(assets, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "assets_seed.json"
    csv_path = output_dir / "assets_seed.csv"

    json_path.write_text(json.dumps(assets, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(assets)

    print(f"Wrote {len(assets)} assets")
    print(json_path)
    print(csv_path)


def main():
    if len(sys.argv) != 3:
        print("Usage: python tools/import_excel_to_seed.py <source.xlsx> <output_dir>")
        raise SystemExit(2)
    assets = read_assets(Path(sys.argv[1]))
    write_outputs(assets, Path(sys.argv[2]))


if __name__ == "__main__":
    main()
